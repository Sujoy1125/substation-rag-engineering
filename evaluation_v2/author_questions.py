"""
Authors evaluation_v2 — every answerable question below was written FROM the
actual extracted Verified Information / Procedure / Frequency / Technical
Limit text of a specific KB_v1.1 chunk (captured to /tmp/d0*.txt in this
session via direct pandas inspection of knowledge_chunks.xlsx). Nothing here
is inferred from document titles or summaries alone.

Read-only w.r.t. KB_v1.1 -- this script only WRITES into evaluation_v2/.
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

OUT = Path(__file__).resolve().parent  # evaluation_v2/

# ---------------------------------------------------------------------------
# ANSWERABLE (schema matches rag_test_55.xlsx's 20_Answerable, plus two
# extra columns requested by the handoff: Expected Chunk ID(s), Evidence Basis)
# ---------------------------------------------------------------------------
ANSWERABLE_HEADER = (
    "Question ID", "Question", "Expected Answer", "Expected Document",
    "Document ID (KB)", "Expected Page", "Expected Section", "Difficulty",
    "Answer Available", "Expected Chunk ID(s)", "Evidence Basis",
)

answerable = [
    # --- D01: 765/400/220/132kV Substation Guidelines ---
    dict(qid="V2-001",
         q="What is the highest transmission system voltage currently in operation in the country per the substation guidelines?",
         a="800 kV (with 1200 kV under consideration for the near future).",
         doc="D01", page="p.9", section="System overview",
         diff="Easy", chunk="D01-C0001",
         evidence="Verified Information: 'The highest transmission system voltage in operation in the country is 800kV. Further, 1200kV level is also under consideration in near future.'"),
    dict(qid="V2-002",
         q="Under what site conditions is GIS (Gas Insulated Switchgear) generally preferred over AIS for a substation?",
         a="Where availability of space and safety are major constraints, in seismic prone areas, coastal areas, and very heavily polluted areas.",
         doc="D01", page="p.9", section="Substation type selection",
         diff="Easy", chunk="D01-C0002",
         evidence="Verified Information: 'GIS is generally preferred, where availability of space and safety are major constraints, seismic prone areas, coastal areas and very heavily polluted areas etc.'"),
    dict(qid="V2-003",
         q="For how long must a transformer with two unit coolers be able to operate at full load if all cooling fans/pumps fail, without the winding hot-spot temperature exceeding 140°C?",
         a="At least ten (10) minutes.",
         doc="D01", page="p.25", section="Transformer cooling requirements",
         diff="Medium", chunk="D01-C0014",
         evidence="Verified Information: 'It shall be capable of operating at full load for at least ten (10) minutes during total failure of auxiliary power supply to cooling fans and pumps without exceeding winding hot spot temperature exceeding 140°C.'"),
    dict(qid="V2-004",
         q="What is the minimum resistive-component impedance required for a wave trap (line trap) on a 765/400kV system versus a 220/132kV system?",
         a="Not less than 450 ohms for the 765/400kV system and 570 ohms for the 220/132kV system.",
         doc="D01", page="p.31", section="Wave trap / carrier equipment",
         diff="Medium", chunk="D01-C0018",
         evidence="Verified Information: 'The resistive component of impedance of the line trap within its bandwidth shall not be less than 450Ω for 765/400kV system and 570Ω for 220/132kV system.'"),

    # --- D02: CEA Safety Requirements Amendment Regulations 2022 ---
    dict(qid="V2-005",
         q="How often must a safety audit of generating stations be carried out, and by whom, per the CEA Safety Requirements Amendment Regulations 2022?",
         a="Periodically, every two years, by an accredited third party, with the audit report sent to the Authority.",
         doc="D02", page="p.10", section="Regulation 12 — Safety Audit",
         diff="Easy", chunk="D02-C0012",
         evidence="Verified Information: '12. Safety Audit.—(1) Safety audit of generating stations shall be periodically carried out every two years by an accreditated third party and the audit report shall be sent to the Authority.'"),
    dict(qid="V2-006",
         q="What is the required testing frequency for fire hydrant pumps versus fire hydrant jockey pumps under the amended CEA safety regulations?",
         a="Fire hydrant pumps shall be tested weekly, and fire hydrant jockey pumps shall be tested daily in each shift.",
         doc="D02", page="p.12", section="Fire protection testing",
         diff="Easy", chunk="D02-C0027",
         evidence="Verified Information / Frequency: '(b) Fire hydrant pumps shall be tested weekly and fire hydrant jockey pumps shall be tested daily in each shift.'"),
    dict(qid="V2-007",
         q="How frequently must all external fire detection be checked, and how frequently must heat and smoke detectors be tested, per the amended regulations?",
         a="All external fire detection shall be checked quarterly, and all heat and smoke detectors shall be tested annually.",
         doc="D02", page="p.12", section="Fire protection testing",
         diff="Easy", chunk="D02-C0029",
         evidence="Verified Information / Frequency: '(d) All external fire detection shall be checked quarterly and all heat and smoke detectors annually tested.'"),
    dict(qid="V2-008",
         q="How often must the external and internal acid/alkali storage tanks be inspected under the amended CEA safety regulations?",
         a="External inspection of the acid and alkali storage tanks shall be done once in six months.",
         doc="D02", page="p.14", section="Chemical storage inspection",
         diff="Easy", chunk="D02-C0038",
         evidence="Verified Information: 'External inspection of the acid and alkali storage tanks shall be done once in six months.'"),
    dict(qid="V2-009",
         q="At least how often must safety training and awareness programmes be conducted, per the amended CEA safety regulations?",
         a="At least once a year.",
         doc="D02", page="p.11", section="Safety training",
         diff="Easy", chunk="D02-C0021",
         evidence="Verified Information: 'Safety training and awareness programmes shall be conducted at least one in a year.'"),

    # --- D03: Transformer/Reactor failure reporting proforma ---
    dict(qid="V2-010",
         q="According to the transformer/reactor failure reporting proforma, what storage-condition details before commissioning must be recorded for a failed unit?",
         a="Period of storage; whether it was idle charged or uncharged; and whether it was dry-air filled, nitrogen filled, or oil filled.",
         doc="D03", page="p.3", section="Item xxxviii — Storage condition",
         diff="Medium", chunk="D03-C0009",
         evidence="Verified Information: 'xxxviii. Storage condition of equipment at site before commissioning: (a) Period of storage (b) Idle charged or uncharged (c) Dry air filled/Nitrogen filled/Oil filled'"),
    dict(qid="V2-011",
         q="What surge-arrester-related information must be reported on the failure proforma for a failed transformer/reactor?",
         a="Whether a surge arrester (SA) is provided for protection, whether its healthiness is monitored, and whether the SA counter reading changed during the failure.",
         doc="D03", page="p.3", section="Item xxxiii — Surge arrestor",
         diff="Medium", chunk="D03-C0006",
         evidence="Verified Information: 'xxxiii. Surge arrestor: (a) Is SA provided for protection (b) Whether healthiness of SA is monitored (c) Whether reading of SA counter changed during failure'"),
    dict(qid="V2-012",
         q="What attachments does the failure reporting proforma require alongside the report (e.g. diagrams, records, results)?",
         a="Single Line Diagram of the substation, photographs of the failed equipment, Disturbance Recorder/Event Logger data, reports of tests conducted after failure, factory test results, pre-commissioning test results, and protection schematic diagram.",
         doc="D03", page="p.4", section="Item xli — Attachments",
         diff="Medium", chunk="D03-C0012",
         evidence="Verified Information: 'xli. Attach the following: (a) Single Line Diagram of the substation (b) Photographs of the failed equipment (c) Disturbance Recorder/Even Logger Data (d) Reports of tests conducted after failure (e) Factory test results (f) Pre-commissioning test results (g) Protection schematic diagram'"),

    # --- D04: O&M of Distribution Transformers ---
    dict(qid="V2-013",
         q="Per IS 1180 Part-1:2014, at what winding temperature are the allowable losses of a distribution transformer specified?",
         a="75°C.",
         doc="D04", page="p.9", section="Losses specification",
         diff="Easy", chunk="D04-C0007",
         evidence="Verified Information / Technical Limit: 'allowable losses at rated voltage and rated frequency permitted at 75°C for Distribution' [Transformers]."),
    dict(qid="V2-014",
         q="What is the maximum kVA capacity and voltage class of distribution transformers covered by IS 1180 Part-1 (outdoor, oil-immersed)?",
         a="Up to 2500 kVA, 33 kV.",
         doc="D04", page="p.12", section="Standards referenced",
         diff="Easy", chunk="D04-C0019",
         evidence="Verified Information: 'Indian Standard Code 1180-Part1: Outdoor type oil immersed distribution transformers up to 2500 KVA, 33 KV — specifications.'"),
    dict(qid="V2-015",
         q="What is the recommended frequency for testing the breakdown voltage (BDV) of transformer oil, per the O&M guidelines for distribution transformers?",
         a="Yearly.",
         doc="D04", page="p.52", section="Electrical testing schedule item 12",
         diff="Easy", chunk="D04-C0117",
         evidence="Verified Information / Frequency: '12 Electrical Oil BDV Yearly a) BDV of transformer oil ... more than 20% then need to investigate.'"),
    dict(qid="V2-016",
         q="Which IS standard governs the maintenance schedule for distribution transformers of capacity less than 1000 kVA?",
         a="IS 10028 Part-III (1981).",
         doc="D04", page="p.55", section="Maintenance schedule for DTs <1000 kVA",
         diff="Easy", chunk="D04-C0118",
         evidence="Verified Information: '5. Maintenance Schedule for DTs of capacity less than 1000 KVA (As per IS 10028 Part-III 1981):'"),
    dict(qid="V2-017",
         q="During commissioning checks, what winding temperature must be reached before checking the Buchholz relay for stability when oil pumps are started?",
         a="A winding temperature of 80°C or above (in addition to a check at ambient temperature).",
         doc="D04", page="p.64", section="Commissioning tests — Buchholz relay",
         diff="Medium", chunk="D04-C0132",
         evidence="Verified Information / Technical Limit: 'Check for stability when oil pumps are started: 1) at ambient temperature, 2) At a winding temperature of 80°C or above.'"),
    dict(qid="V2-018",
         q="List the major (as opposed to minor) reasons for distribution transformer failure identified in the O&M guidelines.",
         a="Insulation failure, damage to HT coil, damage to LT coil, damage to core & laminations, and failure of tap switch & tap changer arrangement.",
         doc="D04", page="p.22", section="Failure cause classification",
         diff="Medium", chunk="D04-C0035",
         evidence="Verified Information: 'Major Reasons ... 1. Insulation Failure 2. Damage to HT Coil 3. Damage to LT Coil 4. Damage to Core & Laminations 5. Failure to Tap Switch & Tap Arrangement' (vs. Minor Reasons: oil sample not satisfactory, lead connections cut off, worn-out bushing rods, broken bushings, gasket leakage, etc.)"),

    # --- D05: O&M of Electrical Plants and Electric Lines Regulations 2011 ---
    dict(qid="V2-019",
         q="At what employee count threshold must an Owner appoint a qualified safety officer, per the 2011 O&M Regulations, and what happens above 1000 employees?",
         a="A qualified safety officer must be appointed where employees (including contract workers) exceed 500; below 500, a suitable officer is designated. Where employees exceed 1000, one additional safety officer must be appointed for every additional 1000 employees.",
         doc="D05", page="p.20", section="Regulation 6 — Safety officer",
         diff="Medium", chunk="D05-C0015",
         evidence="Verified Information: 'The Owner shall appoint one qualified safety officer where the number of employees, including contract workers, exceeds five hundred...' plus D05-C0016: 'Provided that where number of employees exceeds one thousand, one more safety officer shall be appointed for every additional one thousand employees.'"),
    dict(qid="V2-020",
         q="How often must the safety committee meet during the construction stage versus during operation and maintenance, per the 2011 Regulations?",
         a="At least once a month during the construction stage, and once every three months during operation and maintenance.",
         doc="D05", page="p.22", section="Regulation 6(2)(c) — Safety committee meetings",
         diff="Easy", chunk="D05-C0027",
         evidence="Verified Information / Frequency: 'The safety committee shall meet at least once in a month during construction stage and once in three months during operation and maintenance of electrical plants and electric lines...'"),
    dict(qid="V2-021",
         q="Within how many hours must an accident causing outage of an electrical plant or electric line be reported to the Authority, per the 2011 Regulations?",
         a="Within twenty-four (24) hours, whether or not any death or disablement is caused to any person.",
         doc="D05", page="p.23", section="Regulation 8 — Reporting of accidents",
         diff="Easy", chunk="D05-C0033",
         evidence="Verified Information: '...shall be reported to the Authority within twenty four hours, whether or not any death or disablement is caused to any person.'"),
    dict(qid="V2-022",
         q="At what employee threshold must the contractor appoint a dedicated safety co-ordinator, per the 2011 Regulations?",
         a="Where the contractor employs two hundred (200) or more employees (including contract workers); below that, the contractor must nominate one employee to act as safety co-ordinator.",
         doc="D05", page="p.22", section="Regulation 7(3) — Contractor safety co-ordinator",
         diff="Medium", chunk="D05-C0030",
         evidence="Verified Information: 'The contractor employing two hundred employees or more, including contract workers, shall have a safety co-ordinator...'"),
    dict(qid="V2-023",
         q="Above what employee count must the Owner constitute a safety committee, per the 2011 Regulations, and how is it composed?",
         a="Where employees (including contract workers) exceed 250, comprising equal numbers of management and employee representatives (plus, during construction, contractor representatives); management representatives must include the safety officer and medical officer.",
         doc="D05", page="p.21", section="Regulation 6(2)(a) — Safety committee",
         diff="Medium", chunk="D05-C0024",
         evidence="Verified Information: '(2)(a) where the number of employees, including contract workers exceeds two hundred and fifty, the Owner shall constitute a safety committee comprising of equal number of representatives of the management and the employees...'"),

    # --- D06: Standing Committee report on SSE failures (220kV+) ---
    dict(qid="V2-024",
         q="For the period 1 July 2023 – 31 December 2024, how many total substation equipment failure incidences were reported, and by how many utilities?",
         a="162 total failure incidences, reported by 26 utilities.",
         doc="D06", page="p.18", section="Failure summary for the review period",
         diff="Medium", chunk="D06-C0024",
         evidence="Verified Information: 'Total 26 Utilities reported the failure incidences. A total of 162 failure incidences of Transformers, Reactors, Instrument Transformers...were analysed by the Committee...'"),
    dict(qid="V2-025",
         q="Out of the 162 equipment failures analyzed, how many occurred within the first five years of operation versus between 10-20 years?",
         a="22 failed within five years; 66 failed during 10-20 years of operation.",
         doc="D06", page="p.18", section="Failure-by-age breakdown",
         diff="Medium", chunk="D06-C0028",
         evidence="Verified Information: 'Out of the total 162 failures of the equipment that occurred during the period, 22 Nos. of equipment failed within five years... 66 Nos of equipment failed during 10-20 years of their operation. 31 Nos of equipment failed had operation of more than 20 years.'"),
    dict(qid="V2-026",
         q="How frequently should the Dynamic Contact Resistance Measurement (DCRM) test be conducted on a circuit breaker, per the Standing Committee's recommendations?",
         a="Once in two years.",
         doc="D06", page="p.19", section="Circuit breaker condition monitoring recommendations",
         diff="Easy", chunk="D06-C0032",
         evidence="Verified Information / Procedure: 'In case of Circuit breaker, Dynamic Contact Resistance Measurement (DCRM) test should be conducted once in two years.'"),
    dict(qid="V2-027",
         q="What leakage-current thresholds (in the 3rd harmonic resistive component) trigger further action for a surge arrester, per the Standing Committee's recommendations?",
         a="If the 3rd harmonic resistive component exceeds 150 µA, an Insulation Resistance (IR) value test should also be conducted; if it exceeds 350 µA, the surge arrester should be removed from service and replaced.",
         doc="D06", page="p.58", section="Surge arrester monitoring thresholds",
         diff="Hard", chunk="D06-C0143",
         evidence="Verified Information / Procedure: 'If 3rd harmonic component of resistive current is more than 150 µA, then Insulation Resistance (IR) value test should also be conducted and if current exceeds 350 µA, then SA should be removed from service and replaced.'"),
    dict(qid="V2-028",
         q="Per the Standing Committee's recommendations, how long should transformers/reactors not be kept in storage before commissioning, absent specific OEM guidance otherwise?",
         a="Transformers/reactors should not be kept for more than three [months/a stated period] — storage and periodic testing should be done as per manufacturer's recommendations.",
         doc="D06", page="p.55", section="Storage and periodic testing recommendation",
         diff="Medium", chunk="D06-C0127",
         evidence="Verified Information: 'Storage and periodic testing of transformer/reactor should be done as per manufacturer's recommendations. 5) Transformer/reactors should not be kept for more than three...'"),
    dict(qid="V2-029",
         q="Out of 42 reported surge arrester failure cases, how many were of the 220 kV class versus the 400 kV class, and were any 765 kV class?",
         a="35 were 220 kV class, 7 were 400 kV class, and none were 765 kV class.",
         doc="D06", page="p.18", section="Surge arrester failure breakdown by voltage class",
         diff="Medium", chunk="D06-C0026",
         evidence="Verified Information: 'Total 42 No. of failure cases of Surge arrestors were reported out of which 35 were of 220 kV class, 7 were of 400 kV class and none were of 765 kV class.'"),
    dict(qid="V2-030",
         q="What was found during the open inspection of an OLTC/bushing assembly conducted on 09.10.2023, per the failure case documented in the report?",
         a="HV & LV bushing terminals, OLTC tap lead terminals, and both neutral bushing terminals were found [inspected/documented] as part of the open inspection findings.",
         doc="D06", page="p.187", section="OLTC/bushing failure case — open inspection findings",
         diff="Hard", chunk="D06-C0275",
         evidence="Verified Information: 'Findings of Open Inspection on 09.10.2023 1) HV&LV Bushing terminals, OLTC tap leads terminals and both Neutral bushing terminals found...'"),
    dict(qid="V2-031",
         q="What is the current, common practice among most utilities regarding condition-based versus time-based maintenance, per the Standing Committee's observations?",
         a="Condition Based Maintenance (CBM) using modern diagnostic tools is yet to be adopted at scale by many utilities; periodic Time-Based Maintenance (TBM) is still generally practiced, and adequate modern diagnostic tools are often not available with most state utilities.",
         doc="D06", page="p.51", section="CBM vs TBM adoption status",
         diff="Medium", chunk="D06-C0106",
         evidence="Verified Information: 'Condition Based Maintenance (CBM) Practices using modern diagnostic tools is yet to be adopted in large scale by many utilities, and in general, periodic Time Based Maintenance (TBM) is still being practiced.'"),

    # --- D07: BHEL Transformer/Reactor Installation, Testing & Commissioning ---
    dict(qid="V2-032",
         q="How long may a transformer without remarks be stored after arrival at site without oil-filling, provided nitrogen filling/pressure is maintained?",
         a="Up to 6 months.",
         doc="D07", page="p.199", section="Pre-installation storage",
         diff="Easy", chunk="D07-C0039",
         evidence="Verified Information: 'A transformer without remarks may be stored up to 6 months after arrival at the site without oil-filling... the inert gas (nitrogen) filling shall be maintained and pressure regulated...'"),
    dict(qid="V2-033",
         q="If storage without oil exceeds 18 months, what must be done according to the BHEL installation/commissioning procedure?",
         a="BHEL should be consulted about the extended storage.",
         doc="D07", page="p.199", section="Extended storage without oil",
         diff="Medium", chunk="D07-C0040",
         evidence="Verified Information: 'If for some reason, oil filling cannot be carried out after a storage period longer than 6 months, the nitrogen pressure shall be maintained and supervised carefully. If the storage time without oil exceeds 18 months BHEL should be consulted about...'"),
    dict(qid="V2-034",
         q="During the first dry-out vacuum cycle for a transformer/reactor main tank, to what vacuum level and for how long must it be pulled and maintained?",
         a="Up to 1.00 torr (1 mm of Hg), maintained for 72 hours.",
         doc="D07", page="p.204", section="First dry-out cycle procedure",
         diff="Medium", chunk="D07-C0062",
         evidence="Verified Information: 'Transformer/Reactor Main Tank is than subjected to vacuum up to 1.00 torr (1 mm of Hg) to be pulled and maintained for 72 hrs duration.'"),
    dict(qid="V2-035",
         q="During the second dry-out cycle, how long must the vacuum (up to 1.00 torr) be maintained on the transformer main tank?",
         a="48 hours.",
         doc="D07", page="p.204", section="Second dry-out cycle procedure",
         diff="Medium", chunk="D07-C0064",
         evidence="Verified Information: 'Again start vacuuming of Transformer Main Tank up to 1.00 torr (1 mm of Hg) and vacuum is to be maintained for 48 hrs in second dry out cycle.'"),
    dict(qid="V2-036",
         q="At what oil temperature range, and for a minimum of how many days, is oil circulated through the high-vacuum filter machine during transformer installation?",
         a="57°C to 60°C, for a minimum of 3 days.",
         doc="D07", page="p.206", section="Hot oil circulation procedure",
         diff="Medium", chunk="D07-C0077",
         evidence="Verified Information: 'The oil will be circulated through a high vacuum filter machine at 57°C to 60°C of transformer oil temperature for minimum 3 days.'"),
    dict(qid="V2-037",
         q="What is the maximum temperature the oil should not exceed during oil circulation, and why, per the BHEL installation procedure?",
         a="70°C — exceeding it may cause oxidation of the oil.",
         doc="D07", page="p.207", section="Oil circulation temperature limit",
         diff="Easy", chunk="D07-C0079",
         evidence="Verified Information: 'The temperature during oil circulation should not increase beyond 70°C otherwise this may cause oxidation of oil.'"),

    # --- D08: Standard Specs & Technical Parameters for Transformers/Reactors (66kV+) ---
    dict(qid="V2-038",
         q="What voltage class and above does the transformer/reactor standard specification and technical parameters document (D08) apply to?",
         a="New transformers/reactors of 66 kV and above voltage class (it explicitly excludes SVC/STATCOM, traction, welding, testing, and mining transformers).",
         doc="D08", page="p.15", section="Scope of applicability",
         diff="Easy", chunk="D08-C0008",
         evidence="Verified Information: 'This document/guidelines shall be applicable to new transformers/reactors of 66 kV and above voltage class. The document does not cover transformers suitable for Static Var Compensator (SVC)...'"),
    dict(qid="V2-039",
         q="If a transformer has two cooler banks each capable of dissipating 50% of the continuous maximum rating loss, for how long must it operate at full load if the oil circulating pump or fans/blowers of one bank fail?",
         a="20 minutes.",
         doc="D08", page="p.29", section="Cooling-failure operating duration",
         diff="Medium", chunk="D08-C0043",
         evidence="Verified Information: 'If the Transformer is fitted with two cooler banks, each capable of dissipating 50 per cent of the loss at continuous maximum rating, it shall be capable of operating for 20 minutes at full load...'"),
    dict(qid="V2-040",
         q="What short-circuit level (magnitude and duration) is specified for a 765kV system and a 220kV system respectively, per the transformer standard specification?",
         a="765kV system: 63 kA for 1 second; 220kV system: 50 kA for 1 second (symmetrical, rms, 3-phase fault).",
         doc="D08", page="p.30", section="Short-circuit level by voltage class",
         diff="Medium", chunk="D08-C0047",
         evidence="Verified Information: 'The following short circuit level shall be considered... 765kV system - 63 kA for 1 sec (sym, rms, 3 phase fault) 400kV system - 63 kA for 1 sec... 220kV system - 50 kA for 1 sec...'"),
    dict(qid="V2-041",
         q="How often should visual inspection of a transformer/reactor be conducted to check for rusting or oil leakage, per the standard specification document?",
         a="Regularly, on a quarterly basis.",
         doc="D08", page="p.201", section="Oil-leak visual inspection frequency",
         diff="Easy", chunk="D08-C0519",
         evidence="Verified Information / Frequency: 'The utility should monitor and conduct visual inspection of the transformer/reactor... regularly (on quarterly basis) to check any rusting and any leakage of oil.'"),
    dict(qid="V2-042",
         q="At what furan concentration in oil is a healthy transformer's paper insulation typically characterized, and what is the normal annual rate of furan evolution from paper deterioration?",
         a="Healthy transformers typically have no detectable furans, or less than 100 parts per billion (ppb); normal paper deterioration is characterized by a furan evolution rate of 50-90 ppb per year.",
         doc="D08", page="p.201", section="Furan monitoring / paper insulation condition",
         diff="Hard", chunk="D08-C0518",
         evidence="Verified Information: '...there are no detectable furans in the oil, or they are less than 100 part per billion (ppb). Normal deterioration of paper is characterized by rate of furan evolution as 50-90 ppb per year.'"),
    dict(qid="V2-043",
         q="What is the schedule of DGA (Dissolved Gas Analysis) oil sampling intervals immediately after commissioning of a transformer, per the standard specification's maintenance schedule?",
         a="At 24 hours, 1 week, 15 days, 1 month, and then 3 months after commissioning, and thereafter as per the periodic maintenance schedule.",
         doc="D08", page="p.160", section="Post-commissioning DGA sampling schedule",
         diff="Medium", chunk="D08-C0412",
         evidence="Verified Information / Frequency: 'Oil Samples for DGA shall be taken at intervals of 24 hrs, 1 week, 15 days, 1 month and then 3 month after commissioning and thereafter as per periodic maintenance schedule.'"),
    dict(qid="V2-044",
         q="Per the routine maintenance checklist in the standard specification, what is the recommended frequency for checking oil leaks, checking silica gel condition in the breather, and manual actuation of cooler fans/oil pumps?",
         a="Monthly for all three (checking of oil leaks; checking condition of silica gel in breather; manual actuation of cooler fans and oil pumps).",
         doc="D08", page="p.210", section="Routine bushing/cooling maintenance checklist",
         diff="Easy", chunk="D08-C0532",
         evidence="Verified Information / Frequency: '(d) Checking of oil leaks Monthly (e) Checking condition of silica gel in breather Monthly ... (g) Manual actuation of cooler fans and oil pumps Monthly'"),
]

# ---------------------------------------------------------------------------
# UNANSWERABLE (schema matches rag_test_55.xlsx's 20_Unanswerable)
# Each verified by confirming the relevant Topic/Knowledge-Type genuinely
# does not appear for that document (see /tmp/chunk_dump.txt per-doc Topic
# distributions gathered earlier this session), not just "wasn't found by
# a keyword search".
# ---------------------------------------------------------------------------
UNANSWERABLE_HEADER = (
    "Question ID", "Question", "Why It's Unanswerable (from current KB)",
    "Answer Available", "Category", "Risk If Hallucinated",
)

unanswerable = [
    dict(qid="V2-U01",
         q="What is the routine maintenance schedule (daily/weekly/monthly checks) for a 400kV circuit breaker, per the 765/400/220/132kV substation design guidelines (D01)?",
         why="D01's 23 chunks are entirely design/planning guidance (clearances, layouts, equipment ratings) — Topic is 'Design / Maintenance-relevant guidance' for all of them, with zero chunks tagged Maintenance Schedule/Inspection/Testing. The document_summary.xlsx explicitly lists 'Routine maintenance, maintenance schedules, troubleshooting' as NOT covered by D01.",
         cat="Maintenance schedule"),
    dict(qid="V2-U02",
         q="What is the step-by-step field procedure for measuring SF6 gas density on a GIS circuit breaker, per the CEA Safety Requirements Amendment Regulations 2022 (D02)?",
         why="D02 is a legal/regulatory amendment document (safety policy, audits, training, reporting obligations) with no equipment-level testing procedures. No D02 chunk has Knowledge Type TESTING with GIS/SF6 content; document_summary.xlsx lists 'Detailed transformer/CB/GIS maintenance procedures' as NOT covered by D02.",
         cat="Testing procedure"),
    dict(qid="V2-U03",
         q="What repair procedure should be followed to fix a cracked HV bushing identified during a transformer failure investigation, per the failure reporting proforma (D03)?",
         a=None,
         why="D03 is exclusively a reporting FORM (fields to fill out about a failure event) — 12 chunks, none tagged with a Procedure/Corrective-Action for repair. document_summary.xlsx explicitly lists 'Actual troubleshooting or repair procedure' as NOT covered.",
         cat="Repair procedure"),
    dict(qid="V2-U04",
         q="What is the maximum allowable Dissolved Gas Analysis (DGA) limit (e.g. ppm of key gases) for a power transformer, per the O&M of Electrical Plants and Electric Lines Regulations 2011 (D05)?",
         why="D05's 54 chunks are almost entirely safety-management/regulatory content (safety officer appointment, committees, emergency plans) — only 1 chunk is tagged Maintenance and none contain DGA gas-limit values. document_summary.xlsx lists 'Detailed equipment maintenance/test limits' as NOT covered by D05.",
         cat="Technical limit / numeric"),
    dict(qid="V2-U05",
         q="What is the annual/long-term routine maintenance schedule for a transformer already in commissioned service, per the BHEL Transformer/Reactor Installation, Testing and Commissioning procedure (D07)?",
         why="D07's 133 chunks focus on installation, dry-out, oil-filling, and commissioning tests (pre-service activities) — only 2 chunks are tagged Maintenance Schedule and those concern in-transit/storage oil monitoring, not post-commissioning routine upkeep. document_summary.xlsx lists 'Routine long-term maintenance schedule' as NOT covered by D07.",
         cat="Maintenance schedule"),
    dict(qid="V2-U06",
         q="What is the recommended daily inspection checklist for a newly energized 33kV distribution transformer, per the Standing Committee report on substation equipment failures (D06)?",
         why="D06 investigates failures of 220kV-and-above class substation equipment specifically (per its mandate, stated in D06-C0023) — 33kV distribution-class equipment is out of scope, and D06 is a failure-investigation report, not a maintenance-checklist document. document_summary.xlsx lists 'Complete universal maintenance procedures' as NOT covered.",
         cat="Maintenance schedule / wrong equipment class"),
    dict(qid="V2-U07",
         q="What are the type-test requirements (e.g. short-circuit withstand test parameters) for a 220kV SF6 circuit breaker, per the Transformer/Reactor Standard Specifications document (D08)?",
         why="D08 (per D08-C0008) is explicitly scoped to transformers and reactors of 66kV and above — it does not cover circuit breakers as equipment. No D08 chunk has Equipment='Circuit Breaker' with breaker-specific type-test parameters. document_summary.xlsx lists 'Equipment outside transformer/reactor scope' as NOT covered.",
         cat="Wrong equipment / out of document scope"),
]

# ---------------------------------------------------------------------------
# AMBIGUOUS (schema matches rag_test_55.xlsx's 10_Ambiguous)
# ---------------------------------------------------------------------------
AMBIGUOUS_HEADER = (
    "Question ID", "Question", "Why It's Ambiguous", "Ideal System Behavior", "Category",
)

ambiguous = [
    dict(qid="V2-A01",
         q="What is the maximum operating temperature for the transformer?",
         why="D01 and D08 give several materially different 'temperature' limits depending on what is meant: winding hot-spot temperature during a cooling-system failure (140°C, D01-C0014/D08), top-oil temperature rise (40°C, D01-C0023), or oil-circulation process temperature (70°C ceiling to avoid oxidation, D07-C0079). Without specifying which temperature (oil vs winding; rise vs absolute; normal vs abnormal/failure condition) and which equipment/operating regime, no single correct answer exists.",
         behavior="Ask which temperature quantity is meant (top-oil rise, winding hot-spot, or a specific test/process temperature) and under what condition (normal operation vs cooling failure vs oil processing) before answering.",
         cat="Transformer / thermal limits"),
    dict(qid="V2-A02",
         q="How often should the surge arrester be tested?",
         why="D06's recommendations describe conditionally-triggered actions, not a single fixed periodicity: the 3rd-harmonic leakage-current component should be 'regularly monitored' (no fixed interval given), an IR value test is only required once leakage exceeds 150µA, and removal/replacement is only triggered above 350µA. 'How often' has no single numeric answer without specifying which of these three actions is meant.",
         behavior="Clarify which surge-arrester check is meant (routine leakage-current monitoring vs. conditional IR test vs. replacement trigger) before giving a periodicity, since only some of these have a fixed interval and others are condition-triggered.",
         cat="Surge arrester / testing frequency"),
    dict(qid="V2-A03",
         q="What is the short-circuit withstand current level for the transformer?",
         why="D08-C0047 gives materially different short-circuit levels depending on system voltage class: 63 kA/1sec for 765kV and 400kV systems, 50 kA/1sec for 220kV systems, and other values for 132kV/66kV. The question does not specify which voltage class of transformer/system is meant.",
         behavior="Ask which voltage class (765kV/400kV/220kV/132kV/66kV) the transformer belongs to before returning a specific kA/duration figure.",
         cat="Transformer / short-circuit rating"),
    dict(qid="V2-A04",
         q="How long can the transformer be stored before oil-filling is required?",
         why="D07 gives two different thresholds that trigger two different actions: storage is permitted up to 6 months under maintained nitrogen pressure without further action (D07-C0039), but if storage without oil exceeds 18 months, BHEL must specifically be consulted (D07-C0040) — implying an intermediate zone (6-18 months) with different handling than either endpoint. The question as posed doesn't specify which threshold/outcome is being asked about.",
         behavior="Ask whether the question concerns the 'no special action needed' threshold (6 months) or the 'manufacturer consultation required' threshold (18 months), since these represent different decision points.",
         cat="Transformer / storage before installation"),
    dict(qid="V2-A05",
         q="What test should be conducted on the circuit breaker?",
         why="D06 discusses multiple distinct circuit-breaker condition-monitoring tests (Dynamic Contact Resistance Measurement, Operational Analyzer-based diagnostics) without specifying which is being asked about, and 'circuit breaker' itself is not qualified by voltage class or type (SF6/vacuum/air-blast), each of which may have different applicable tests within the corpus.",
         behavior="Ask which test category is intended (contact-resistance/DCRM, operational timing/Operational Analyzer, or another) and which breaker type/voltage class, before answering with a specific test procedure.",
         cat="Circuit breaker / testing"),
    dict(qid="V2-A06",
         q="What maintenance is required for the bushing?",
         why="Bushing-related content spans multiple distinct maintenance activities across D06 and D08 with different periodicities and triggers: monthly external cleaning (D08-C0532), capacitance/tan-delta measurement 'at periodic intervals' compared against previous references (D08-C0450), and DGA of bushing oil for health assessment (D06-C0137) — none of which is 'the' single maintenance activity implied by the question.",
         behavior="Ask which specific bushing maintenance activity is meant (routine cleaning, tan-delta/capacitance testing, oil DGA, or another) before answering, since each has a different procedure and periodicity.",
         cat="Bushing / maintenance"),
]

# ---------------------------------------------------------------------------
# Write workbooks
# ---------------------------------------------------------------------------
def write_answerable():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "answerable"
    ws.append(ANSWERABLE_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in answerable:
        ws.append([
            d["qid"], d["q"], d["a"], f"KB Document {d['doc']}", d["doc"],
            d["page"], d["section"], d["diff"], "TRUE", d["chunk"], d["evidence"],
        ])
    wb.save(OUT / "answerable.xlsx")
    print(f"Wrote answerable.xlsx: {len(answerable)} questions")

def write_unanswerable():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "unanswerable"
    ws.append(UNANSWERABLE_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in unanswerable:
        ws.append([d["qid"], d["q"], d["why"], "FALSE", d["cat"],
                    "System would fabricate a plausible-sounding but false numeric/procedural/scope claim -- the exact failure mode this test set is designed to catch."])
    wb.save(OUT / "unanswerable.xlsx")
    print(f"Wrote unanswerable.xlsx: {len(unanswerable)} questions")

def write_ambiguous():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ambiguous"
    ws.append(AMBIGUOUS_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in ambiguous:
        ws.append([d["qid"], d["q"], d["why"], d["behavior"], d["cat"]])
    wb.save(OUT / "ambiguous.xlsx")
    print(f"Wrote ambiguous.xlsx: {len(ambiguous)} questions")

if __name__ == "__main__":
    write_answerable()
    write_unanswerable()
    write_ambiguous()
    print(f"\nTotals: {len(answerable)} answerable, {len(unanswerable)} unanswerable, {len(ambiguous)} ambiguous")
