"""End-to-end generation evaluation over all 57 evaluation_v2 questions.

STEPS 6-8 of the engineering plan: score the generation layer on answerable
(must answer), unanswerable (must abstain) and ambiguous (must ask) questions,
separately, and report the safety/coverage pair the confidence-gating claim
will eventually rest on.

Modes
-----
    (default)          Retrieval-side only. No LLM, no API key. Reports how
                       often the gold evidence reaches the context at all —
                       the ceiling on everything downstream.

    --live             Full pipeline on every question. Deterministic scoring:
                       citations, abstention, clarification. Answer
                       correctness is NOT scored in this mode.

    --live --judge     Adds the LLM judge on attempted answers. The judge is
                       uncalibrated; its numbers mean nothing without the
                       agreement figure from --agreement below.

    --review-sheet     Writes an xlsx of every attempted answer next to its
                       reference answer, with an empty "Human Verdict" column.
                       Grade a sample by hand, save, then:

    --agreement FILE   Reads the graded sheet and reports Cohen's kappa
                       between the judge and the human grades.

Usage
-----
    python experiments/run_generation_eval.py                     # dry, free
    python experiments/run_generation_eval.py --live
    python experiments/run_generation_eval.py --live --judge --review-sheet
    python experiments/run_generation_eval.py --agreement experiments/review_sheet_graded.xlsx
    python experiments/run_generation_eval.py --live --limit 3    # cheap trial

Nothing here writes to the KB or to any evaluation workbook.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

import openpyxl

from src.evaluation.eval_loader import load_all
from src.confidence.gate import ConfidenceGate, ConfidenceModel, UncalibratedGateError
from src.confidence.gated import run_gate, split_results, summarise_changes
from src.evaluation.generation_eval import (
    JudgeVerdict,
    build_report,
    compare_reports,
    compute_agreement,
    print_report,
    score_ambiguous,
    score_answerable,
    score_unanswerable,
)
from src.confidence.signals import extract_signals
from src.evaluation.judge import AnswerJudge, judge_all
from src.generation.answer import AnswerStatus
from src.generation.llm import LLMUnavailableError, client_from_env
from src.generation.pipeline import DEFAULT_TOP_K, RAGPipeline, load_kb
from src.retrieval.benchmark import page_numbers

REVIEW_HEADER = [
    "Question ID",
    "Difficulty",
    "Question",
    "Reference Answer (gold)",
    "System Status",
    "System Answer",
    "Citations",
    "Cited Chunk IDs",
    "Gold Chunk IDs",
    "Gold Location Cited",
    "Judge Verdict",
    "Judge Reason",
    "Human Verdict",  # <- fill in: CORRECT / PARTIALLY_CORRECT / INCORRECT
    "Human Notes",
]

VALID_HUMAN_VERDICTS = {
    JudgeVerdict.CORRECT.value,
    JudgeVerdict.PARTIALLY_CORRECT.value,
    JudgeVerdict.INCORRECT.value,
}


def parse_args():
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("--live", action="store_true", help="call the configured LLM")
    p.add_argument("--judge", action="store_true", help="also run the LLM judge (implies --live)")
    p.add_argument(
        "--gated",
        action="store_true",
        help=(
            "apply the calibrated confidence gate and report ungated vs gated "
            "side by side (implies --live). This is STEP 10 — run it on --split "
            "holdout, once. Requires a fitted model from calibrate_confidence.py"
        ),
    )
    p.add_argument("--review-sheet", action="store_true", help="write an xlsx for manual grading")
    p.add_argument("--agreement", metavar="XLSX", help="compute judge/human agreement from a graded sheet")
    p.add_argument("--top-k", type=int, default=DEFAULT_TOP_K)
    p.add_argument("--limit", type=int, default=None, help="first N questions per class (cheap trials)")
    p.add_argument("--label", default="", help="label for this run in the report")
    p.add_argument(
        "--split",
        choices=["all", "calibration", "holdout"],
        default="all",
        help=(
            "which frozen split to evaluate (evaluation_v2/split_v1.json). "
            "'all' for baseline measurement where nothing is being tuned; "
            "'calibration' while fitting the confidence gate; "
            "'holdout' ONCE, to report the final result"
        ),
    )
    return p.parse_args()


def apply_split(args, answerable, unanswerable, ambiguous):
    """Filter the three classes to the requested side of the frozen split."""
    if args.split == "all":
        return answerable, unanswerable, ambiguous

    from src.evaluation.splits import load_split

    split = load_split()
    keep = split.ids_for(args.split)
    answerable = [q for q in answerable if q.question_id in keep]
    unanswerable = [q for q in unanswerable if q.question_id in keep]
    ambiguous = [q for q in ambiguous if q.question_id in keep]

    print(f"SPLIT: {args.split} — {split.summary()}")
    if args.split == "holdout":
        print(
            "\n  You are spending the holdout. These questions are the only\n"
            "  out-of-sample evidence you have that the gate generalises rather\n"
            "  than being fitted to the calibration set. Tuning anything after\n"
            "  reading this run turns them into calibration data, and the final\n"
            "  number stops being defensible.\n"
        )
    return answerable, unanswerable, ambiguous


# ---------------------------------------------------------------------------
# agreement mode
# ---------------------------------------------------------------------------


def run_agreement(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    try:
        j_idx = header.index("Judge Verdict")
        h_idx = header.index("Human Verdict")
        q_idx = header.index("Question ID")
    except ValueError:
        print(f"{path}: expected columns 'Question ID', 'Judge Verdict', 'Human Verdict'.")
        return 1

    pairs, skipped, invalid = [], 0, []
    for row in rows:
        if row is None or row[q_idx] is None:
            continue
        judge = str(row[j_idx] or "").strip().upper()
        human = str(row[h_idx] or "").strip().upper()
        if not human:
            skipped += 1
            continue
        if human not in VALID_HUMAN_VERDICTS:
            invalid.append((str(row[q_idx]), human))
            continue
        if judge not in VALID_HUMAN_VERDICTS:
            skipped += 1
            continue
        pairs.append((judge, human))

    if invalid:
        print(f"Ignored {len(invalid)} rows with an unrecognised Human Verdict:")
        for qid, v in invalid[:10]:
            print(f"  {qid}: {v!r} (expected one of {sorted(VALID_HUMAN_VERDICTS)})")

    if not pairs:
        print("No graded rows found. Fill in the 'Human Verdict' column and re-run.")
        return 1

    ag = compute_agreement(pairs)
    print(f"\nJudge vs human agreement over {ag.n} graded answers "
          f"({skipped} ungraded rows skipped)")
    print(f"  raw agreement   {ag.raw_agreement:.3f}")
    print(f"  Cohen's kappa   {ag.cohens_kappa:.3f}  ({ag.verdict()})")
    print("\n  confusion (rows = judge, cols = human)")
    labels = sorted(ag.confusion)
    print("    " + " ".join(f"{l[:9]:>10}" for l in [""] + labels))
    for a in labels:
        print(f"    {a[:9]:>10} " + " ".join(f"{ag.confusion[a][b]:>10}" for b in labels))
    print(
        "\n  Report this kappa alongside every judged number. A judge whose\n"
        "  agreement with you is unknown is not a measurement."
    )
    return 0


# ---------------------------------------------------------------------------
# dry mode
# ---------------------------------------------------------------------------


def run_dry(answerable, unanswerable, ambiguous, chunks, top_k) -> int:
    from src.generation.context import build_context
    from src.retrieval.equipment_aware_v2 import EquipmentAwareRetrieverV2
    from src.retrieval.retrievers import BM25Retriever

    retriever = EquipmentAwareRetrieverV2(BM25Retriever())
    retriever.index(chunks)

    print("MODE: dry — retrieval only, no LLM call, no API key required.\n")

    hits, ranks, misses = 0, [], []
    for q in answerable:
        results = retriever.retrieve(q.question, top_k=top_k)
        gold = set(q.expected_chunk_ids)
        # Same matching rule as score_answerable(): exact chunk id first, then
        # the benchmark's page-level definition. Using a stricter rule here
        # than the scorer uses would report a ceiling the scorer never applies.
        rank = next(
            (i for i, r in enumerate(results, start=1) if r.chunk.chunk_id in gold), None
        )
        if rank is None:
            rank = next(
                (
                    i
                    for i, r in enumerate(results, start=1)
                    if r.chunk.document_id == q.gold.expected_document_id
                    and page_numbers(r.chunk.pdf_page) & page_numbers(q.gold.expected_page)
                ),
                None,
            )
        if rank:
            hits += 1
            ranks.append(rank)
        else:
            misses.append(q.question_id)

    n = len(answerable)
    print(f"ANSWERABLE ({n}) — gold chunk inside the top-{top_k} context")
    print(f"  reached the model  {hits}/{n} = {hits / n:.3f}")
    if ranks:
        print(f"  mean rank when present  {sum(ranks) / len(ranks):.2f}")
    print(f"  never reached the model  {misses}")
    print(
        "\n  This is the ceiling on answer correctness: the model cannot ground an\n"
        "  answer in evidence it was never shown. Questions listed above can only\n"
        "  be answered correctly by accident."
    )

    print(f"\nUNANSWERABLE ({len(unanswerable)}) / AMBIGUOUS ({len(ambiguous)}) — context size")
    for label, qs in (("unanswerable", unanswerable), ("ambiguous", ambiguous)):
        sizes, docs = [], []
        for q in qs:
            ctx = build_context(retriever.retrieve(q.question, top_k=top_k))
            sizes.append(len(ctx.text))
            docs.append(len(ctx.document_ids()))
        if sizes:
            print(
                f"  {label:<13} mean evidence {sum(sizes) // len(sizes)} chars, "
                f"mean {sum(docs) / len(docs):.1f} documents"
            )
    print(
        "\n  These questions always retrieve something — BM25 always returns its top\n"
        "  matches. Abstention therefore has to be a judgement about whether the\n"
        "  evidence answers the question, not about whether evidence exists."
    )
    print("\nDry run complete. No generation behaviour has been measured.")
    return 0


# ---------------------------------------------------------------------------
# live mode
# ---------------------------------------------------------------------------


def write_review_sheet(path: Path, answerable, results, scores, judged) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "review"
    ws.append(REVIEW_HEADER)

    by_id = {s.question_id: s for s in scores}
    for q, r in zip(answerable, results):
        s = by_id[q.question_id]
        if not s.answered:
            continue  # nothing to grade
        j = judged.get(q.question_id)
        ws.append(
            [
                q.question_id,
                q.gold.difficulty,
                q.question,
                q.gold.expected_answer,
                s.status,
                r.answer.answer_text,
                "; ".join(c.short() for c in r.answer.citations),
                "; ".join(c.chunk_id for c in r.answer.citations),
                "; ".join(q.expected_chunk_ids),
                "YES" if s.page_level_cited else "NO",
                j.verdict.value if j else JudgeVerdict.NOT_JUDGED.value,
                j.reason if j else "",
                "",  # Human Verdict — to fill in
                "",  # Human Notes
            ]
        )

    widths = [12, 10, 55, 55, 20, 70, 26, 26, 20, 18, 18, 55, 18, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def run_live(answerable, unanswerable, ambiguous, chunks, args) -> int:
    try:
        client = client_from_env()
    except LLMUnavailableError as e:
        print(f"Cannot run live: {e}")
        print("Set OPENAI_API_KEY in .env, or run without --live for the free dry mode.")
        return 2

    print(f"MODE: live — provider={client.provider} model={client.model} top_k={args.top_k}")
    if args.judge:
        print("Judge: ENABLED. Judge verdicts are uncalibrated — grade a sample and")
        print("       run --agreement before quoting any judged number.\n")
    else:
        print("Judge: disabled. Answer correctness will NOT be measured.\n")

    pipeline = RAGPipeline(chunks, llm=client, top_k=args.top_k).index()

    # Stop early rather than burning 57 API calls against a dead connection.
    # The first run of this harness failed all 9 calls with "Connection error"
    # and still printed a full metrics table; fail fast and say why instead.
    MAX_CONSECUTIVE_ERRORS = 3

    class ConnectionDead(RuntimeError):
        pass

    state = {"consecutive_errors": 0}

    def run_class(name, questions):
        out = []
        for i, q in enumerate(questions, start=1):
            result = pipeline.answer(q.question)
            status = result.answer.status
            line = f"  [{name} {i}/{len(questions)}] {q.question_id} -> {status.value}"
            if result.error:
                line += f"   {result.error}"
            print(line)
            out.append(result)

            if status is AnswerStatus.LLM_ERROR:
                state["consecutive_errors"] += 1
                if state["consecutive_errors"] >= MAX_CONSECUTIVE_ERRORS:
                    raise ConnectionDead(result.error or "the model could not be reached")
            else:
                state["consecutive_errors"] = 0
        return out

    def abort(reason: str) -> int:
        print(f"\n{'!' * 72}")
        print(f"ABORTED after {MAX_CONSECUTIVE_ERRORS} consecutive failures to reach the model.")
        print(f"  {reason}")
        print("!" * 72)
        print(
            "\nNothing was measured, so no report is produced. Common causes:\n"
            "  - no internet route to api.openai.com (campus/corporate firewall,\n"
            "    proxy required, or captive portal)\n"
            "  - the API key is revoked, mistyped, or has no credit\n"
            "  - OPENAI_BASE_URL is set to something unreachable\n"
            "\nQuick check:\n"
            '  python -c "import openai,os;'
            'print(openai.OpenAI(api_key=os.getenv(\'OPENAI_API_KEY\')).models.list().data[0].id)"\n'
            "\nThe free dry mode still works and needs no network:\n"
            "  python experiments/run_generation_eval.py"
        )
        return 3

    try:
        print("Running answerable...")
        a_results = run_class("A", answerable)
        print("Running unanswerable...")
        u_results = run_class("U", unanswerable)
        print("Running ambiguous...")
        m_results = run_class("M", ambiguous)
    except ConnectionDead as e:
        return abort(str(e))

    a_scores = [score_answerable(q, r) for q, r in zip(answerable, a_results)]
    u_scores = [score_unanswerable(q, r) for q, r in zip(unanswerable, u_results)]
    m_scores = [score_ambiguous(q, r) for q, r in zip(ambiguous, m_results)]

    # --- confidence gate (Step 10) -------------------------------------
    # Both systems are scored by the SAME scorers over the same runs; the only
    # variable is whether the gate rewrote the status. Anything else would
    # confound the comparison with a difference in measurement.
    gated_report = None
    gate_changes = None
    gated_records = None
    if args.gated:
        try:
            model = ConfidenceModel.load()
        except FileNotFoundError as e:
            print(f"\nCannot run gated: {e}")
            return 4
        if not model.is_calibrated:
            print(
                "\nCannot run gated: the confidence model has no fitted weights.\n"
                "Fit it on the calibration split first:\n"
                "  python experiments/run_generation_eval.py --live --split calibration\n"
                "  python experiments/calibrate_confidence.py --from <json> --max-unsafe-rate <policy>"
            )
            return 4

        print(f"\nGate: {model.fitted_on}, fitted on {model.fitted_n_questions} questions")
        print(
            f"      answer_threshold={model.answer_threshold:.2f} "
            f"clarify_threshold={model.clarify_threshold:.2f}"
        )
        gate = ConfidenceGate(model)
        try:
            a_gated = run_gate(gate, a_results)
            u_gated = run_gate(gate, u_results)
            m_gated = run_gate(gate, m_results)
        except UncalibratedGateError as e:
            print(f"\nCannot run gated: {e}")
            return 4

        all_gated = a_gated + u_gated + m_gated
        gate_changes = summarise_changes(all_gated)
        gated_records = [g.to_dict() for g in all_gated]

        gated_report = build_report(
            f"{args.label or 'gated'} [GATED]",
            [score_answerable(q, g.gated) for q, g in zip(answerable, a_gated)],
            [score_unanswerable(q, g.gated) for q, g in zip(unanswerable, u_gated)],
            [score_ambiguous(q, g.gated) for q, g in zip(ambiguous, m_gated)],
        )

    judged = {}
    if args.judge:
        print("\nJudging attempted answers...")
        judged = judge_all(AnswerJudge(client), answerable, a_results)
        by_id = {s.question_id: s for s in a_scores}
        for qid, jr in judged.items():
            if qid in by_id and jr.verdict is not JudgeVerdict.NOT_JUDGED:
                by_id[qid].judge_verdict = jr.verdict.value
                by_id[qid].judge_reason = jr.reason
        errors = [j for j in judged.values() if j.error]
        if errors:
            print(f"  {len(errors)} judge call(s) failed: {errors[0].error}")

    label = args.label or f"{client.provider}/{client.model} top_k={args.top_k} split={args.split}"
    report = build_report(label, a_scores, u_scores, m_scores)
    print_report(report)

    if gated_report is not None:
        print_report(gated_report)
        print("\n--- what the gate did ---")
        for transition, n in gate_changes.items():
            marker = "" if transition.split(" -> ")[0] == transition.split(" -> ")[1] else "  <-- changed"
            print(f"  {n:>3}  {transition}{marker}")
        if all(t.split(" -> ")[0] == t.split(" -> ")[1] for t in gate_changes):
            print(
                "\n  The gate changed nothing. Either the thresholds are too low for\n"
                "  this data, or the signals do not separate it. Either way there is\n"
                "  no gating effect to report."
            )
        compare_reports(report, gated_report)
        print(
            "\n  The claim is the PAIR: unsafe assertions down, coverage held.\n"
            "  A gate that drives unsafe to zero by abstaining from everything has\n"
            "  not improved the system, and the coverage column will show it."
        )

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_json = REPO_ROOT / "experiments" / f"generation_eval_{stamp}.json"
    out_json.write_text(
        json.dumps(
            {
                "experiment": "generation_eval",
                "timestamp_utc": datetime.now(timezone.utc).isoformat(),
                "provider": client.provider,
                "model": client.model,
                "top_k": args.top_k,
                "judge_enabled": args.judge,
                "judge_calibrated": False,
                "gated": args.gated,
                "gated_report": None if gated_report is None else gated_report.to_dict(),
                "gate_transitions": gate_changes,
                "gate_decisions": gated_records,
                "note": (
                    "Judge verdicts are uncalibrated. Report Cohen's kappa from "
                    "--agreement alongside any judged rate."
                ),
                "split": args.split,
                "report": report.to_dict(),
                # Confidence signals are computed here, at the point the run
                # happened, so the calibrator never has to reconstruct them
                # from a serialised result and risk drifting from signals.py.
                "confidence_signals": {
                    q.question_id: extract_signals(r).to_dict()
                    for q, r in list(zip(answerable, a_results))
                    + list(zip(unanswerable, u_results))
                    + list(zip(ambiguous, m_results))
                },
                "per_question": {
                    "answerable": [r.to_dict() for r in a_results],
                    "unanswerable": [r.to_dict() for r in u_results],
                    "ambiguous": [r.to_dict() for r in m_results],
                },
                "judge_results": {
                    qid: {
                        "verdict": j.verdict.value,
                        "reason": j.reason,
                        "disputed_facts": j.disputed_facts,
                        "error": j.error,
                    }
                    for qid, j in judged.items()
                },
            },
            indent=2,
        )
    )
    print(f"\nSaved: {out_json.relative_to(REPO_ROOT)}")

    if args.review_sheet:
        out_xlsx = REPO_ROOT / "experiments" / f"review_sheet_{stamp}.xlsx"
        write_review_sheet(out_xlsx, answerable, a_results, a_scores, judged)
        print(f"Saved: {out_xlsx.relative_to(REPO_ROOT)}")
        print(
            "\nNext: fill the 'Human Verdict' column (CORRECT / PARTIALLY_CORRECT /\n"
            "INCORRECT) for a sample, save, then run:\n"
            f"  python experiments/run_generation_eval.py --agreement {out_xlsx.relative_to(REPO_ROOT)}"
        )
    return 0


def main() -> int:
    args = parse_args()

    if args.agreement:
        return run_agreement(args.agreement)

    answerable, unanswerable, ambiguous = load_all()
    answerable, unanswerable, ambiguous = apply_split(args, answerable, unanswerable, ambiguous)
    if args.limit:
        answerable = answerable[: args.limit]
        unanswerable = unanswerable[: args.limit]
        ambiguous = ambiguous[: args.limit]

    chunks = load_kb()
    print(
        f"KB: {len(chunks)} chunks | answerable {len(answerable)} | "
        f"unanswerable {len(unanswerable)} | ambiguous {len(ambiguous)}\n"
    )

    if args.judge or args.gated:
        args.live = True
    if args.live:
        return run_live(answerable, unanswerable, ambiguous, chunks, args)
    return run_dry(answerable, unanswerable, ambiguous, chunks, args.top_k)


if __name__ == "__main__":
    raise SystemExit(main())
