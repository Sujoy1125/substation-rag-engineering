"""
KB_v1 Cross-File Validation Script
Run this any time the knowledge base, equipment inventory, catalog, or evaluation
files change, to catch consistency regressions before they reach the RAG pipeline.

Usage: python3 validate_kb.py
Exit code 0 = PASS, 1 = FAIL (see printed report for details)
"""
import openpyxl, re, sys
from collections import Counter

BASE = "."  # adjust to point at the KB_v1 folder when running standalone

def load(path, sheet=None, data_only=True):
    wb = openpyxl.load_workbook(path, data_only=data_only)
    return wb[sheet] if sheet else wb.active

def nonblank_rows(ws, min_row=2):
    return [r for r in ws.iter_rows(min_row=min_row, values_only=True) if r[0]]

def main():
    errors = []
    warnings = []

    # --- Load core files ---
    cat_ws = load(f"{BASE}/document_catalog.xlsx")
    cat_rows = nonblank_rows(cat_ws)
    cat_ids = set(r[0] for r in cat_rows)

    chunk_ws = load(f"{BASE}/knowledge_chunks.xlsx", "knowledge_chunks")
    chunk_rows = nonblank_rows(chunk_ws)
    chunk_ids = [r[0] for r in chunk_rows]

    # --- Check 1: duplicate chunk IDs ---
    dupes = [k for k, v in Counter(chunk_ids).items() if v > 1]
    if dupes:
        errors.append(f"Duplicate chunk IDs: {dupes}")

    # --- Check 2: every chunk's source ID exists in catalog ---
    orphan_sources = set(r[1] for r in chunk_rows if r[1]) - cat_ids
    if orphan_sources:
        errors.append(f"Chunks reference source IDs not in document_catalog.xlsx: {orphan_sources}")

    # --- Check 3: every chunk has real content (not empty) ---
    empty_content = [r[0] for r in chunk_rows if not (r[11] or r[12])]
    if empty_content:
        errors.append(f"Chunks with no Verified Information AND no Procedure text: {empty_content}")

    # --- Check 4: PDF Page field format sanity ---
    bad_pages = [r[0] for r in chunk_rows if not re.match(r'^PDF p\.? \d', str(r[18] or ''))]
    if bad_pages:
        errors.append(f"Chunks with malformed PDF Page field: {bad_pages}")

    # --- Check 5: catalog completeness (URL + authority level) ---
    missing_url = [r[0] for r in cat_rows if not r[4]]
    missing_auth = [r[0] for r in cat_rows if not r[10]]
    if missing_url:
        errors.append(f"Catalog entries missing source URL: {missing_url}")
    if missing_auth:
        errors.append(f"Catalog entries missing authority level: {missing_auth}")

    # --- Check 6: equipment inventory chunk/doc references valid ---
    try:
        eq_ws = load(f"{BASE}/equipment_inventory.xlsx", "Equipment Inventory")
        eq_header = [c.value for c in eq_ws[1]]
        chunk_col = eq_header.index("D09 Chunk / Page") if "D09 Chunk / Page" in eq_header else None
        doc_col = eq_header.index("Relevant Maintenance Documents") if "Relevant Maintenance Documents" in eq_header else None
        chunk_id_set = set(chunk_ids)
        bad_chunk_refs, bad_doc_refs = [], []
        for row in eq_ws.iter_rows(min_row=2, values_only=True):
            if chunk_col is not None and row[chunk_col]:
                for cid in re.findall(r'D\d{2}-C\d{4}', str(row[chunk_col])):
                    if cid not in chunk_id_set:
                        bad_chunk_refs.append((row[0], cid))
            if doc_col is not None and row[doc_col]:
                for did in [x.strip() for x in str(row[doc_col]).split(";")]:
                    if did and did not in cat_ids:
                        bad_doc_refs.append((row[0], did))
        if bad_chunk_refs:
            errors.append(f"Equipment inventory refs nonexistent chunk IDs: {bad_chunk_refs}")
        if bad_doc_refs:
            errors.append(f"Equipment inventory refs nonexistent document IDs: {bad_doc_refs}")
    except FileNotFoundError:
        warnings.append("equipment_inventory.xlsx not found, skipped check 6")

    # --- Check 7: answerable-question evidence coverage ---
    try:
        ev_ws = load(f"{BASE}/rag_test_55.xlsx", "20_Answerable")
        d09_chunks = [r for r in chunk_rows if str(r[0]).startswith("D09")]
        def covers(page_str, target):
            parts = [int(p) for p in page_str.replace("PDF p. ", "").split("-")]
            lo, hi = parts[0], parts[-1]
            return lo - 1 <= target <= hi + 1
        no_evidence = []
        for row in ev_ws.iter_rows(min_row=2, values_only=True):
            qid, page_field = row[0], row[5]
            target = int(re.search(r'\d+', str(page_field).split("-")[0]).group())
            if not any(covers(c[18], target) for c in d09_chunks):
                no_evidence.append((qid, page_field))
        if no_evidence:
            errors.append(f"Answerable questions with no covering chunk: {no_evidence}")
    except FileNotFoundError:
        warnings.append("rag_test_55.xlsx not found, skipped check 7")

    # --- Report ---
    print("=" * 60)
    print(f"VALIDATION REPORT — {len(errors)} error(s), {len(warnings)} warning(s)")
    print("=" * 60)
    for e in errors:
        print("  [FAIL]", e)
    for w in warnings:
        print("  [WARN]", w)
    if not errors:
        print("  PASS — no consistency errors found.")
    return 0 if not errors else 1

if __name__ == "__main__":
    sys.exit(main())
